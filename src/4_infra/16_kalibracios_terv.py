"""16_kalibracios_terv.py — kalibrációs munkalap a 91 paraméterhez.

MIÉRT. A `parameterek.csv` azt tartja nyilván, hogy egy paraméter MENNYIRE
megbízható (kategória, státusz, forrás). Ez a munkalap arra válaszol, hogy
MIT KELL CSINÁLNI VELE: mit jelent, mire hat, milyen képlettel számolható,
milyen adat kell hozzá, mi az irodalmi sáv, és át lehet-e venni konvencióként.

A tábla NEM helyettesíti a regisztert — abból építkezik. A regiszter a
forrás; ha egy paraméter státusza változik, ott kell javítani.

Kimenet: output/tables/kalibracios_terv.xlsx
Futtatás: python src/4_infra/16_kalibracios_terv.py
"""

from __future__ import annotations

import pathlib

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = next(p for p in pathlib.Path(__file__).resolve().parents
            if (p / "CLAUDE.md").exists())
REG = REPO / "docs" / "regiszter"
KI = REPO / "output" / "tables" / "kalibracios_terv.xlsx"

# ---------------------------------------------------------------- sablonok
# Több paraméter ugyanabba a családba tartozik; a közös mezőket egyszer írjuk le.
JV = ("Bayes-i poszterior, magyar adaton becsült rendszerben (Jakab–Világi, "
      "MNB WP 2008/9). Egyetlen paraméter külön újrabecslése elrontja a "
      "készlet belső konzisztenciáját.")
JV_ADAT = "Negyedéves magyar makro-idősorok (1995–2007) + teljes DSGE-újrabecslés."
JV_MOST = "NEM — a projekt nem futtat Bayes-i becslést; a készlet együttesen becsült."
JV_KELL = "Teljes újrabecslés mai magyar mintán (önálló kutatási projekt, nem fér a decemberi leadásba)."

SOKK = ("AR(1) sokk-perzisztencia. A determinisztikus euró-szcenárióban NEM "
        "aktív (a sokkok nincsenek hajtva), csak sztochasztikus futásban él.")
SOKK_KEPLET = ("AR(1) regresszió a becsült sokkfolyamatra: e_t = rho·e_{t-1} + u_t; "
               "rho = Cov(e_t, e_{t-1}) / Var(e_{t-1}).")

SZARM = "SZÁRMAZTATOTT — nem szabad paraméter, más paraméterekből számolódik."

# oszlopok: nev, jelentes, szerep, hatas, keplet, adatigeny, irodalom, most, kell, konvencio
P: dict[str, tuple] = {}


def add(nev, jelentes, szerep, hatas, keplet, adatigeny, irodalom, most, kell, konv, *names):
    for n in names:
        P[n] = (nev, jelentes, szerep, hatas, keplet, adatigeny, irodalom, most, kell, konv)


# --- háztartás ------------------------------------------------------------
add("Relatív kockázatkerülés", "Az intertemporális helyettesítési rugalmasság reciproka.",
    "A fogyasztási Euler-egyenlet kamatérzékenységét szabja meg.",
    "Az aggregált fogyasztás reakciója a reálkamatra; közvetve az egész euró-csatorna erőssége. Nem típus-specifikus.",
    "Euler-egyenletből: c_t = E[c_{t+1}] − (1/sigma)·(r_t − E[pi_{t+1}]). A sigma a fogyasztásnövekedés és a reálkamat közti regresszió meredekségének reciproka.",
    "Negyedéves reálfogyasztás + ex-ante reálkamat idősor.", "1,0–2,5 (Smets–Wouters 1,38; JV 1,814)",
    JV_MOST, JV_KELL, "IGEN — széles körben átvett tartomány.", "sigma")
add("Szokásformálás", "A múltbeli fogyasztás súlya a jelenlegi hasznosságban.",
    "Inerciát ad a fogyasztásnak: nem ugrik azonnal az új szintre.",
    "A GDP-hatás időbeli lefutása (hump-shape). A hosszú távú szintet nem.",
    "A fogyasztás AR(1) együtthatója a becsült Euler-egyenletben: c_t = h/(1+h)·c_{t-1} + …",
    "Negyedéves reálfogyasztás idősor.", "0,5–0,8 (JV 0,646)", JV_MOST, JV_KELL,
    "IGEN — konvencionális sáv.", "habit")
add("Korlátozott háztartások aránya", "A hitelkorlátos („kézből a szájba élő”) háztartások súlya.",
    "Ez a rész nem optimalizál: elfogyasztja a folyó jövedelmét.",
    "A fogyasztás jövedelemérzékenysége; tartja a keresletet, ha a kamatcsatorna gyenge.",
    "KÖZVETLENÜL MÉRHETŐ: a nettó likvid vagyonnal nem rendelkező háztartások aránya a vagyonfelmérésben.",
    "Háztartási vagyonfelmérés (HFCS), magyar hullámok.", "0,25–0,50 (JV 0,25 = survey-feltevés, nem poszterior)",
    "NEM — de nem is DSGE-becslés kérdése.",
    "HFCS magyar hullám mikroadata. Egy DIREKT mérés jobb, mint bármelyik DSGE-becslés. NYITOTT TÉTEL.",
    "RÉSZBEN — a szám konvencionális, de erre a paraméterre létezik jobb, direkt forrás.", "om_no")
add("Inverz Frisch-rugalmasság", "A munkakínálati görbület a bér-Phillips-görbében.",
    "Megszabja, mennyire reagál a ledolgozott óraszám a reálbérre.",
    "A munkapiaci alkalmazkodás sebessége; a lam_w meredekségén át az inflációs dinamika.",
    "Munkakínálati egyenlet: w_t = fii·l_t + sigma·c_t. A fii a bér–munka regresszió meredeksége.",
    "Negyedéves ledolgozott óra + reálbér.", "1,0–5,0 (makro 2–5; mikro 0,5–1)", JV_MOST, JV_KELL,
    "IGEN — de a makro/mikro szakadék ismert vita.", "fii")
add("Bérmarkup-rugalmasság", "A munkatípusok közti helyettesítési rugalmasság.",
    "A bérfelárat és a lam_w meredekségét szabja.",
    "A bér-Phillips-görbe meredeksége. Nem típus-specifikus.",
    "theta_w = 1/(bérmarkup − 1); pl. 25%-os bérfelár → theta_w ≈ 4. A lam_w inputja.",
    "Bérfelár-becslés vagy irodalmi konvenció.", "3,0–6,0 (20–50% bérfelár)", JV_MOST, JV_KELL,
    "IGEN — tipikus konvenció.", "theta_w")

# --- ár- és bérragadósság -------------------------------------------------
add("Calvo ár-ragadósság", "Annak valószínűsége, hogy egy cég egy negyedévben NEM változtat árat.",
    "Az ár-Phillips-görbe meredekségét (lam_p) állítja elő.",
    "Az infláció perzisztenciája és a reálhatás nagysága. Mindhárom típus ugyanazt kapja, de eltérő mc_j mellett relatívár-különbséget formál.",
    "lam_p = (1−xi_p)(1−beta·xi_p)/xi_p. MIKROADATBÓL KÖZVETLENÜL: az átlagos árváltoztatási gyakoriság = 1/(1−xi_p) negyedév.",
    "Fogyasztói/termelői ár-mikroadat (egyedi árjegyzések gyakorisága).",
    "0,66–0,92 makróbecslés; MIKROADAT 0,4–0,7 (gyakoribb árváltoztatás). ⚠ A JV 0,921 ≈ 12,7 negyedév — a mikroadat ezt ritkán támasztja alá.",
    "NEM.", "MNB/KSH ár-mikroadat (árjegyzés-szintű gyakoriság).",
    "IGEN, de ⚠ a makro- és mikrobecslés szisztematikusan eltér.", "xi_p")
add("Ár-indexáció", "A múltbeli inflációhoz való automatikus igazítás súlya.",
    "Backward-looking tagot ad a Phillips-görbének.",
    "Az infláció tehetetlensége. Nem típus-specifikus.",
    "A Phillips-görbe múltbeli inflációs tagjának együtthatója: vth_p/(1+beta·vth_p).",
    "Negyedéves inflációs idősor.", "0,0–0,5 (JV 0,431)", JV_MOST, JV_KELL, "IGEN.", "vth_p")
add("Calvo export-ár ragadósság ⚫ HALOTT", "Örökölt exportár-merevségi paraméter.",
    "⚫ A v09 model blokkjában NEM SZEREPEL — csak a szintén halott lam_x-et számolja.",
    "SEMMIRE. Ha valaki bekötné, a t54 őr elbukik.",
    "lam_x = (1−xi_x)(1−beta·xi_x)/xi_x — kiszámolódik, de sehol nem használt.",
    "—", "0,66–0,92 (elvben)", "TÁRGYTALAN — halott paraméter.",
    "Előbb el kell dönteni, kell-e külön exportár-Phillips-görbe a v09-be.",
    "TÁRGYTALAN.", "xi_x")
add("Export-ár indexáció ⚫ HALOTT", "Örökölt exportár-indexációs paraméter.",
    "⚫ A v09-ben SEHOL nem szerepel.", "SEMMIRE.", "—", "—", "0,0–0,5 (elvben)",
    "TÁRGYTALAN — halott paraméter.", "Lásd xi_x.", "TÁRGYTALAN.", "vth_x")
add("Calvo bér-ragadósság", "Annak valószínűsége, hogy egy negyedévben nem alkalmazkodik a bér.",
    "A bér-Phillips-görbe meredekségét (lam_w) állítja elő.",
    "A reálbér alkalmazkodásának sebessége, közvetve a foglalkoztatás és a határköltség.",
    "lam_w = (1−xi_w)(1−beta·xi_w)/(xi_w·(1+theta_w·fii)). Mikroadatból: az átlagos bérváltoztatási gyakoriság.",
    "Cégszintű vagy egyéni bér-mikroadat.",
    "0,6–0,85 (JV 0,657). ⚠ Saját mérésünk (A21) GYENGE nominális merevséget talált 2023–24-ben: a cégek 10,1%-a CSÖKKENTETTE a nominálbért.",
    "RÉSZBEN — az Opten-panel bérdinamikát ad, de csak egy év-pár (2023/24) és magas inflációs környezetben.",
    "Hosszabb cégszintű bérpanel, normál inflációs időszakot is lefedve.",
    "IGEN, de ⚠ a saját magyar adat feszültségben van vele (A21).", "xi_w")
add("Bér-indexáció", "A múltbeli inflációhoz igazodó automatikus béremelés súlya.",
    "Backward-looking tag a bér-Phillips-görbében.", "A bérinfláció tehetetlensége.",
    "A bér-egyenlet múltbeli inflációs együtthatója.", "Bér- és inflációs idősor.",
    "0,0–0,5 (JV 0,185)", JV_MOST, JV_KELL, "IGEN.", "vth_w")

# --- monetáris és zárás ---------------------------------------------------
add("Kamatsimítás", "A jegybanki alapkamat AR(1) tagja a Taylor-szabályban.",
    "Fokozatossá teszi a kamatpálya-alkalmazkodást.",
    "A monetáris transzmisszió időzítése a LEBEGŐ rezsimben. Az euró-rezsimben (uni=1) kikapcsol.",
    "Taylor-szabály becslése: r_t = gam_i·r_{t-1} + (1−gam_i)·phi_pi·pi_t + …",
    "Alapkamat + infláció negyedéves idősor.", "0,7–0,9 (JV 0,761)", JV_MOST, JV_KELL, "IGEN.", "gam_i")
add("Taylor inflációs együttható", "A jegybank reakciója az inflációs eltérésre.",
    "A Taylor-elv teljesülését biztosítja (>1), ami a determinációhoz kell.",
    "A nominális horgony a lebegő rezsimben. Az euró-rezsimben kikapcsol.",
    "Ugyanabból a Taylor-regresszióból, mint a gam_i.", "Alapkamat + infláció.",
    "1,2–2,0 (Taylor eredeti 1,5; JV 1,379). ⚠ >1 KÖTELEZŐ a determinációhoz.",
    JV_MOST, JV_KELL, "IGEN.", "phi_pi")
add("NFA-zárási paraméter", "Az adósságérzékeny kamatprémium meredeksége.",
    "Technikai: stacionáriussá teszi a nettó külföldi pozíciót egy kis nyitott gazdaságban.",
    "Nélküle a modell egységgyököt tartalmazna. Gazdaságilag kicsi hatás.",
    "Schmitt-Grohé–Uribe (2003) zárás: prémium = nu_b·(bstar − bstar_ss). Rendszerint NEM becsült, hanem kicsire állított.",
    "—", "0,0001–0,01 (technikai; minél kisebb, annál kevésbé torzít)",
    "⚠ KÉTÉRTELMŰ: a regiszter JV-poszteriorként vette át, a modellben viszont technikai zárásként működik. Tisztázandó.",
    "A pontos JV-tábla ellenőrzése: becsült érték vagy kalibrációs zárás?",
    "IGEN — technikai konvenció.", "nu_b")
add("Uniózárás", "NFA-zárás a monetáris uniós rezsim kamategyenletében.",
    "Ugyanaz, mint a nu_b, de az uni=1 ágon.",
    "A reálárfolyam és az NFA hosszú távú viselkedése az euró-rezsimben. ⚠ Egy korábbi hibás érték (0,01) +34,7%-os reálárfolyamot adott.",
    "Technikai kalibráció, nem becslés. Őrrel kezelendő.", "—",
    "0,01–0,5 (technikai)", "TÁRGYTALAN — technikai zárás.",
    "Érzékenységi scan (-DNUUNI), nem adat.", "IGEN — technikai konvenció.", "nu_uni")
add("Szuverén felár átgyűrűzése a hazai kamatba",
    "A szuverén kockázati felár hányad része jelenik meg a hazai kockázatmentes kamatban.",
    "A kamat- és UIP-zárásban: r − zsov·sov.",
    "Az egész euró-hatás nagyságrendjének egy része: mennyi jut le a szuverén felár csökkenéséből az általános kamatszintre. Nem típus-specifikus.",
    "REGRESSZIÓ: Δ(BUBOR − EURIBOR) = zsov·Δ(magyar szuverén felár) + kontrollok. A meredekség maga a zsov.",
    "Két NYILVÁNOS idősor: magyar szuverén CDS/EMBI felár + BUBOR–EURIBOR különbözet (v. határidős prémium).",
    "Nincs bevett konvenció; 0,3–0,7 plauzibilis.",
    "✅ IGEN — ez a legolcsóbb horgony a horgonyzatlanok közül. Tisztán makró-idősor, nem szenved a programvezéreltségtől és az egy pre-periódustól.",
    "Semmi új: a két idősor nyilvános. Fél nap.",
    "NEM — magyar-specifikus, mérni kell.", "zsov")
add("Diszkontfaktor", "A jövőbeli hasznosság negyedéves leszámítolása.",
    "Minden előretekintő egyenletben (Euler, Phillips, NFA).",
    "A steady-state reálkamat: r_ss = 1/beta − 1. ⚠ NEM származtatott: ő az INPUTJA a lam_p/lam_w-nek.",
    "beta = 1/(1 + r_reál_negyedéves). A 0,99 ≈ 4%/év reálkamatnak felel meg.",
    "Hosszú távú reálkamat.", "0,985–0,995 (gyakorlatilag mindig 0,99)",
    "MEGVAN — a .mod fixen 0,99.", "—",
    "IGEN — ez a legkonvencionálisabb paraméter az egész modellben.", "beta")

# --- termelés -------------------------------------------------------------
add("Tőkehányad a határköltségben", "A tőkeköltség súlya a típus határköltségében.",
    "mc_j = zeta_j·rk_j + (1−zeta_j)·wz_j − a. A tőkeköltség-érzékenységet szabja.",
    "🟠 Mennyire érzékeny a típus a tőkeköltségre — tehát az egész euró-csatornára. A BK-valid dekompozíció szerint a teljes technológiai heterogenitás kivétele a küszöböt csak ~3%-kal mozdítja.",
    "zeta_j = tőkejövedelem / hozzáadott érték = (ÉCS + üzemi eredmény) / (ÉCS + üzemi eredmény + személyi jellegű ráfordítás), típusonként.",
    "Cégszintű eredménykimutatás (ÉCS, üzemi eredmény, bérköltség) VAGY KSH ágazati tőkejövedelem-hányad.",
    "0,30–0,40 aggregált tőkehányad; a JV szektorértékei (0,14/0,17) ennél jóval alacsonyabbak, mert a MŰKÖDÉSI határköltségre vonatkoznak.",
    "✅ RÉSZBEN — az Opten-panel tartalmazza a szükséges sorokat. A nehézség nem az adat, hanem az E/D/L leképezés.",
    "Az aggregált KSH-adat megvan; a méret × exportorientáció szerinti bontáshoz a cégpanelből kell újraszámolni.",
    "NEM — jelenleg a JV szektorértékeinek ÁTVITELE, nem becslés. A .mod maga írja: „Ez ATVITEL, nem becsles.”",
    "zeta_E", "zeta_D", "zeta_L")
add("Munka aránya a munka+import kompozitban",
    "A bérköltség súlya a munka és az importált köztes input kompozitjában.",
    "wz_j = aa_j·w + (1−aa_j)·rer. Az (1−aa_j) az import-/árfolyamkitettség.",
    "🟠 A magyar duális szerkezet központi paramétere: az exportáló KKV költségének 55%-a importált input, a hazaiénak csak 20%-a. Ezért az árfolyam a két típust másképp érinti.",
    "aa_j = személyi jellegű ráfordítás / (személyi jellegű ráfordítás + importált anyagjellegű ráfordítás), típusonként.",
    "Cégszintű bérköltség + anyagjellegű ráfordítás, az import-tartalom elkülönítésével.",
    "Nincs bevett irodalmi érték — ország- és szektorspecifikus.",
    "⚠ RÉSZBEN — a bérköltség megvan a panelben, de az anyagjellegű ráfordítás IMPORT-tartalma nem közvetlenül.",
    "Cégszintű import-adat (vám-/VIES-adat) vagy ágazati import-tartalom (TiVA) hozzákapcsolása.",
    "NEM — átvitel a JV szektoraiból.", "aa_E", "aa_D", "aa_L")
add("Tőke-kompozit helyettesítés", "Helyettesítési rugalmasság a tényezőkeresletben (tőke).",
    "A tőkekereslet reakciója a relatív tényezőárakra.",
    "🟡 Közös minden típusra, de a típusonként eltérő zeta_j hatását alakítja.",
    "CES tényezőkeresletből: a tényezőarány log-változása / a relatív ár log-változása.",
    "Ágazati tényezőár- és tényezőmennyiség-idősor.", "0,5–1,0 (Cobb–Douglas = 1)",
    JV_MOST, JV_KELL, "IGEN.", "rho_kz")
add("Munka–import helyettesítés", "Helyettesítési rugalmasság a munka és az import között.",
    "Az import és a munka közti szubsztitúció a kompozitban.",
    "🟡 Közös, de az eltérő aa_j hatását alakítja.",
    "Mint a rho_kz, a munka–import arányra.", "Bér- és importár-idősor + mennyiségek.",
    "0,3–1,0", JV_MOST, JV_KELL, "IGEN.", "rho_z")
add("Negyedéves amortizációs ráta", "A tőkeállomány negyedéves leírási üteme.",
    "k_j = (1−delta)·k_j(−1) + delta·i_j.",
    "🟡 A tőkedinamika sebessége. Közös paraméter, nem hoz létre méretkülönbséget.",
    "delta = éves écs / bruttó tárgyi eszköz, negyedévesítve (osztva 4-gyel vagy 1−(1−delta_év)^(1/4)).",
    "Cégszintű vagy nemzeti számlás écs + tőkeállomány.",
    "0,020–0,030 negyedéves (8–12%/év)",
    "✅ MEGVAN — saját mérés: 0,0242 (Opten-panel), az átvett 0,025-höz nagyon közel. Két független forrás egyezik (A09).",
    "—", "IGEN — de nálunk már mérve is van.", "delta")

# --- típussúlyok ----------------------------------------------------------
add("Típus méretsúlya", "A vállalattípus részesedése az aggregált kibocsátásból.",
    "Az aggregált tőke, beruházás és a származtatott súlyok (wd, wx, shm) alapja.",
    "🔵 Az aggregálást és az összetételt mozgatja; az egyedi y_j-t nem közvetlenül.",
    "om_j = a típus árbevétele / az összes vállalat árbevétele.",
    "Cégszintű árbevétel, méret és exportstátusz szerint bontva.",
    "Nincs irodalmi érték — országspecifikus szerkezeti adat.",
    "⚠ FELTÉTELES — megvan (0,2555/0,1844/0,5601), de csak a 10+ fős körön BELÜLI részesedésként; a mikrocégek nincsenek a panelben.",
    "KSH/Eurostat SBS mikrokör-bontás (0–9 fős cégek), vagy az EC/ECB SAFE éves köre, amely mikrocégekre is bont.",
    "NEM — magyar szerkezeti adat.", "om_E", "om_D", "om_L")
add("Exportárbevétel-arány", "A típus árbevételének exportra jutó hányada.",
    "y_j = (1−phi_j)·d_j + phi_j·x_j; a wd_j és wx_j súlyokat is meghatározza.",
    "🔴 KÖZVETLEN piaci-orientációs driver: eldönti, mennyire érinti a típust a reálárfolyam vs a hazai kereslet.",
    "phi_j = exportárbevétel / összes árbevétel, típusonként.",
    "Cégszintű árbevétel export/belföld bontásban.",
    "Nincs irodalmi érték — országspecifikus.",
    "✅ phi_L MEGVAN (0,3649, négy tizedesig egyezik az átvett 0,365-tel). ⚠ phi_E és phi_D DEFINÍCIÓFÜGGŐ: tág def. 0,376/0,000; 25%-os def. 0,691/0,023.",
    "Semmi új adat — a kérdés KUTATÓI DÖNTÉS: hol húzzuk meg az „exportáló KKV” határát. Ezt scanelni kell (10/25/40%).",
    "NEM.", "phi_E", "phi_D", "phi_L")
add("Foglalkoztatási súly", "A típus részesedése az aggregált munkakeresletből.",
    "Az aggregált ll súlyai; a bér–fogyasztás visszacsatolás.",
    "🔵 Az aggregált munkapiaci visszacsatolás; az egyedi l_j-t nem közvetlenül hajtja.",
    "shl_j = a típus foglalkoztatottjai / összes foglalkoztatott.",
    "Cégszintű létszámadat méret és exportstátusz szerint.",
    "Nincs irodalmi érték.",
    "⚠ FELTÉTELES — megvan (0,1566/0,3775/0,4659), de csak a 10+ fős körre.",
    "Lásd om_j: SBS mikrokör-bontás.", "NEM.", "shl_E", "shl_D", "shl_L")

# --- pénzügyi blokk (BGG) -------------------------------------------------
add("BGG külső finanszírozási felár rugalmassága",
    "Mennyivel nő a finanszírozási felár, ha a tőkeáttétel egy egységgel nő.",
    "efp_j = chi_j·(q_j + k_j − nw_j) + tsov_j·sov + tbank_j·bank.",
    "🔴 AZ EGYIK LEGERŐSEBB SZEGMENSDRIVER. A t35 scan szerint a chi-sorrend MEGFORDÍTÁSA megfordítja a szegmenssorrendet (−1,22 pp → +0,74 pp). ⚠ A jelenlegi 0,06/0,06/0,02 pontosan a VISSZAVONT V04 állítás számai (K01 konfliktus).",
    "PANEL-REGRESSZIÓ: hitelfelár_it = chi·tőkeáttétel_it + cég- és időhatások. A chi a meredekség. Elméletileg a BGG-szerződésből: chi = f(monitoring-költség, csődvalószínűség).",
    "Cégszintű HITELFELÁR (nem implicit ráta!) + mérlegpozíció, méret szerint, PIACI árazású időszakban.",
    "0,042–0,067 (Christensen–Dib 2008; BGG-konvenció) — de MÉRET SZERINTI BONTÁS NÉLKÜL. Egyetlen közös chi-t becsül az irodalom.",
    "❌ MEGPRÓBÁLTUK, NEM SIKERÜLT: a magyar panelen chi_S ≈ +0,002 (erősen attenuált alsó korlát), chi_L pedig egyáltalán nem azonosított (n=230, t=−0,78, rossz előjel). Ok: az implicit ráta mérési hibája (A15) és a programvezéreltség (A04).",
    "Cégszintű ÚJ-SZERZŐDÉSES hitelkamat (nem állományi implicit ráta) méret szerint, lehetőleg 2015–2020-as, nem programvezérelt időszakra. MNB hitelregiszter.",
    "⚠ RÉSZBEN — a SZINT átvehető (0,042–0,067), de a MÉRET SZERINTI ARÁNY nem: arra nincs irodalmi alap. Javaslat: szimmetrikus alap + scan.",
    "chi_E", "chi_D", "chi_L")
add("Tőkeáttétel", "A tőkeállomány és a nettó vagyon aránya (k/n).",
    "nw_j = omega·(nw_j(−1) + lev_j·(ret_j − r(−1) + infl)). A hozamrés hatását skálázza.",
    "🔴 A pénzügyi akcelerátor erőssége típusonként: magasabb lev = erősebb felerősítés.",
    "lev_j = medián(eszközök összesen / saját tőke). Keresztpróba: 1/(1 − kötelezettségek/eszközök).",
    "Cégszintű mérleg (eszközök, saját tőke), pozitív saját tőkére szűrve.",
    "k/n ≈ 2,0 (BGG 1999 konvenció); Christensen–Dib 2008 hasonló.",
    "✅ MEGVAN — Opten-panel medián: 1,939 (E) / 1,719 (D) / 2,337 (L). A SORREND (L>E>D) mérőfüggetlen; a SZINT mérőfüggő, ezért sávban közöljük (A07, A08). ⚠ Az OPTEN=0 alapág MÉGIS a cáfolt 1,6=1,6-ot futtatja (K02 konfliktus).",
    "—", "RÉSZBEN — a ~2-es nagyságrend konvencionális, de a méret szerinti sorrend saját mérés.",
    "lev_E", "lev_D", "lev_L")
add("Beruházási kiigazítási költség", "A beruházás megváltoztatásának konvex költsége.",
    "i_j = … + 1/((1+beta)·psi_j)·q_j + omega_acc_j·acc_j. A q hatása 1/psi_j-vel arányos.",
    "🔴 KÖZVETLEN beruházási aszimmetria: alacsonyabb psi = rugalmasabb beruházás. ⚠ A jelenlegi sorrend (KKV 8 < nagyvállalat 13) a KKV-t teszi rugalmasabbá — az irodalom az ELLENKEZŐJÉT sugallja.",
    "A beruházási Euler-egyenletből: a beruházás reakciója a Tobin-Q-ra = 1/((1+beta)·psi). Panel-regresszió: Δi_it = (1/((1+beta)psi))·q_it + …",
    "Cégszintű beruházás + Tobin-Q proxy (piaci/könyv szerinti érték vagy profitabilitás), méret szerint.",
    "2,5–8,0 (CEE 2005 ≈ 2,48; Smets–Wouters ≈ 5,74) — MÉRET SZERINTI BONTÁS NÉLKÜL. ⚠ A rögös beruházás (Khan–Thomas 2008; Bachmann et al.) szerint a KIS cégek beruházása RÖGÖSEBB, ami inkább psi_S > psi_L-t sugallna.",
    "❌ NEM PRÓBÁLTUK. Az Opten-panel tartalmaz beruházási adatot, de Q-proxy nélkül.",
    "Cégszintű beruházási mikrodinamika méret és exportstátusz szerint, Q-proxyval.",
    "⚠ RÉSZBEN — a szint átvehető, a méret szerinti SORREND nem; sőt az irodalom a jelenlegi ellen szól. KÖTELEZŐ SCAN.",
    "psi_E", "psi_D", "psi_L")
add("Tőkeár–hozam időzítési súly", "A q és a tőkehozam relatív súlya a vállalati saját tőkehozamban.",
    "A nettóvagyon-egyenlet hozamtagjának összetétele.",
    "🟡 Közös mindhárom típusra, de a meglévő BGG-aszimmetriákat skálázza.",
    "BGG-konvenció; nem önállóan becsült.", "—", "0,95–0,99 (BGG 1999)",
    "NEM — nem magyar becslés.", "Elvben cégszintű tőkehozam-dekompozíció.",
    "IGEN — bevett BGG-konvenció.", "eps_qw")
add("Nettó vagyon perzisztenciája", "A vállalkozók túlélési/tőkefelhalmozási rátája.",
    "nw_j = omega_nw·(…). A tőkefelhalmozás lassúságát adja.",
    "🟡 A pénzügyi akcelerátor időbeli lefutása; a lev_j különbségekkel EGYÜTT hat.",
    "BGG-konvenció: omega_nw ≈ 1 − (vállalkozói kilépési ráta). 0,95 ≈ 5 év átlagos élettartam.",
    "Cégszintű túlélési ráta.", "0,94–0,99 (BGG 1999: 0,9728)",
    "RÉSZBEN — a cégpanelből a kilépési ráta elvben számolható.",
    "Cég-demográfiai adat (belépés/kilépés) méret szerint.",
    "IGEN — bevett BGG-konvenció.", "omega_nw")

# --- transzmisszió --------------------------------------------------------
add("Szuverén felár átgyűrűzése a vállalati felárba",
    "A szuverén sokk hányad része jelenik meg a típus finanszírozási felárában.",
    "efp_j = … + tsov_j·sov + …",
    "🔴 CSAK DIFFERENCIÁLT ÁGBAN. A TSCEN=3 alapágban mindhárom típus 0,175-öt kap, tehát NEM épít be méretkülönbséget. A TSCEN=1/2 exogén módon KKV- vagy nagyvállalati előnyt ad.",
    "REGRESSZIÓ: Δ(vállalati hitelkamat_j) = tsov_j·Δ(szuverén felár) + tbank_j·Δ(bankközi ráta) + kontrollok, MÉRET SZERINT külön.",
    "Méret szerinti új-szerződéses vállalati hitelkamat + szuverén felár, hosszabb idősoron.",
    "Nincs bevett szint. SZERKEZETI TÁMOGATÁS: Bottero–Lenzu–Mezzanotti (JIE 2018) szerint a szuverén sokk MENNYISÉGI átgyűrűzése méret-semleges, a REÁLHATÁS viszont nem — ez épp a jelenlegi modellszerkezetet igazolja.",
    "❌ MEGPRÓBÁLTUK (t25/t25b): mind a négy specifikációban a NAGYVÁLLALATI átgyűrűzés magasabb, de EGYIK SEM szignifikáns (minden CI tartalmazza a nullát).",
    "MNB új-szerződéses vállalati kamatstatisztika cellaszinten (méret × összeg × fixálás × futamidő + volumenek, 2015–2024).",
    "NEM — és a semleges (TSCEN=3) alapértelmezés a védhető választás.",
    "tsov_E", "tsov_D", "tsov_L")
add("Banki forrásköltség átgyűrűzése a vállalati felárba",
    "A banki sokk hányad része jelenik meg a típus felárában.",
    "efp_j = … + tbank_j·bank.",
    "🔴 CSAK DIFFERENCIÁLT ÁGBAN — az alapág méret-semleges (0,45 mindhárom típusnak).",
    "Mint a tsov_j, a bankközi ráta együtthatója a méret szerinti kamatregresszióban.",
    "Ugyanaz, mint a tsov_j.",
    "⚠ VAN IRODALOM, ÉS ELLENTMOND A SAJÁT MÉRÉSÜNKNEK: Horváth–Kotlebová–Širaňová (JFS 2018, 36:12–21) szerint a kamattranszmisszió CSAK a kisvállalati hitelekre teljes, minden más kategóriára hiányos.",
    "❌ Lásd tsov_j (t25/t25b) — a magyar pontbecslés az ELLENKEZŐ irányba mutat, nem szignifikánsan.",
    "Ugyanaz, mint a tsov_j: MNB új-szerződéses kamatstatisztika.",
    "NEM — az irodalom és a magyar adat ELLENTÉTES irányba mutat (F06), ezért a semleges alapértelmezés a helyes.",
    "tbank_E", "tbank_D", "tbank_L")

# --- access-blokk ---------------------------------------------------------
add("Hitelhozzáférési állapot perzisztenciája",
    "Mennyire ragad be egy szegmens javult hitelhozzáférési állapota.",
    "acc_j = rho_acc·acc_j(−1) − lambda_acc_j·efp_j (csak E és D típusra).",
    "🔴 A hosszú távú access-hatás 1/(1−rho_acc)-kal arányos, ami rho→1 közelében ROBBAN: 0,85→6,7×; 0,9673→30,6×. ⚠ CSAK a KKV-kra hat, mert L-access állapot nincs. ⚠ ACCSCALE=100 mellett rho_acc≥0,93 már terminálisan BK-INVALID.",
    "SZEGMENS-SZINTŰ dinamika: a szegmens hozzáférési arányának AR(1) együtthatója egy azonosított sokk után. NEM a cégszintű státusz-perzisztencia!",
    "Szegmens-szintű hozzáférési idősor + azonosított kamat-/felársokk.",
    "Nincs bevett érték; a DSGE-kben 0,8–0,95 tipikus.",
    "❌ NEM. A 0,9673 egy CÉGSZINTŰ, BINÁRIS státusz-perzisztencia leíró statisztikája — a négy teljes évvel rendelkező cégek 92,4%-a EGYSZER SEM váltott státuszt, tehát a szám főként ÁLLANDÓ cégheterogenitást mér, nem dinamikus alkalmazkodást. Ezért az A11 állítást visszavontuk.",
    "Szegmens-szintű dinamikus becslés, VAGY olyan panelmodell, amely a fix céghatást leválasztja az időbeli alkalmazkodásról (pl. a státuszt VÁLTÓ cégek almintáján).",
    "NEM — és a jelenlegi szám nem is horgony, csak érzékenységi pont.", "rho_acc")
add("Felár → hitelhozzáférés rugalmasság (1. lépcső)",
    "Mennyivel javul a hozzáférés, ha a finanszírozási felár csökken.",
    "acc_j = rho_acc·acc_j(−1) − lambda_acc_j·efp_j.",
    "🔴 ⚠ A modell a lambda_acc-ot és az omega_acc-ot KÜLÖN-KÜLÖN NEM AZONOSÍTJA — csak a SZORZATUKAT (A22, t52e-ben numerikus nullára mérve). L-változat NINCS.",
    "Elvben: IV-regresszió, ahol egy exogén felársokk hatását mérjük a hozzáférési arányra. ⚠ DE: a modellen keresztül csak a lambda·omega szorzat azonosítható, külön egyik sem.",
    "Exogén felársokk + szegmens-szintű hozzáférési arány.",
    "Nincs bevett érték — modellspecifikus konstrukció.",
    "❌ NEM, és ELVILEG SEM a 2021–24-es magyar epizódból: a támogatott programok kiiktatták a kamatciklust a KKV-hozzáférésből (A04/A06). A BUBOR 12,8 pontot mozgott, a hozzáférési arányok legfeljebb 2,2-t.",
    "Nem programvezérelt időszak, VAGY EIBIS/SAFE-típusú felmérés, amely a korlátozottságot és a beruházást együtt méri. ⚠ Még ezekből is CSAK A SZORZAT jön ki.",
    "NEM — átvett érték a v07_access-ből; küszöbformában közlendő.",
    "lambda_acc_E", "lambda_acc_D")
add("Hitelhozzáférés → beruházás rugalmasság (2. lépcső)",
    "Mennyivel nő a beruházás, ha javul a hozzáférés — VÁLTOZATLAN jövedelmezőség mellett.",
    "i_j = … + omega_acc_j·acc_j. Additív forcing tag a beruházási Eulerben.",
    "🔴 ⚠ omega_acc_L = 0 — a nagyvállalatnak DEFINÍCIÓ SZERINT nincs acc-egyenlete. EZ A MODELL LEGNAGYOBB EGYETLEN FELTEVÉSE, és nincs is benne a 91 paraméterben, mert nem érték, hanem HIÁNYZÓ EGYENLET. Soha nem scaneltük.",
    "Elvben: a pénzügyileg korlátozott és nem korlátozott cégek beruházási rátájának különbsége. ⚠ Csak a lambda·omega szorzat azonosítható.",
    "Cégszintű korlátozottsági mutató + beruházási ráta.",
    "IRÁNY-TÁMOGATÁS: a nem korlátozott cégek nettó beruházási rátája ~7,8 százalékponttal magasabb, és a hatás KISEBB a nagyvállalatoknál — ez támogatja az omega_acc_L < omega_acc_S irányt. Az ECB SAFE szerint a nagyvállalati elutasítási arány 1–2%, a KKV-ké 5%.",
    "❌ NEM (lásd lambda_acc). ⚠ CSAPDA: egy programhatás-IV-ből kapott Δberuházás/Δhozzáférés hányadost NEM szabad omega_acc-nak nevezni — az a program teljes finanszírozási hatását viszi, nem a strukturális rugalmasságot.",
    "EIBIS / EC-ECB SAFE magyar bontás (a Bizottság ÉVES köre, amely minden EU-tagállamot lefed — a negyedéves ECB-kör csak euróövezeti!).",
    "NEM — átvett érték; küszöbformában közlendő.",
    "omega_acc_E", "omega_acc_D")

# --- kereslet, GDP-súlyok, vertikális link --------------------------------
for nm, cim, mit in [("sc", "Fogyasztás GDP-aránya", "a fogyasztás"),
                     ("si", "Beruházás GDP-aránya", "a beruházás"),
                     ("sg", "Kormányzati kereslet GDP-aránya", "a kormányzati kereslet")]:
    add(cim, f"{mit.capitalize()} súlya a kiadási oldali GDP-ben.",
        "A GDP-azonosság összeállítása: y = sc·c + si·ii + sg·g + sx·xx − sm·im.",
        "🔵 Főleg az AGGREGÁLT y mérését változtatja; az egyedi y_j-t nem.",
        f"{cim.split()[0]} = {mit} folyó áron / GDP folyó áron, több éves átlag.",
        "KSH nemzeti számlák, negyedéves vagy éves.",
        "Országspecifikus — nincs irodalmi érték, adatból jön.",
        "✅ IGEN — KSH nemzeti számlákból KÖZVETLENÜL frissíthető, semmi trükk.",
        "Semmi új: nyilvános KSH-adat.", "NEM — magyar adat, de triviálisan elérhető.", nm)
add("Export GDP-aránya", "Az export súlya a GDP-ben és a külső pozícióban.",
    "A GDP-azonosságban és a nettó külföldi pozíció (bstar) felhalmozásában.",
    "🔵/🟡 A bstar-on és a monetáris záráson át VISSZA is csatol.",
    "sx = export folyó áron / GDP folyó áron.", "KSH nemzeti számlák.",
    "Országspecifikus.", "✅ IGEN — KSH-ból közvetlenül.", "Semmi új.", "NEM.", "sx")
add("Import GDP-aránya", "Az import súlya a GDP-ben és a külső pozícióban.",
    "Mint az sx, az import oldalon.", "🔵/🟡 Ugyanott.",
    "sm = import folyó áron / GDP folyó áron.", "KSH nemzeti számlák.",
    "Országspecifikus.", "✅ IGEN — KSH-ból közvetlenül.", "Semmi új.", "NEM.", "sm")
add("Exportkereslet perzisztenciája", "A külpiaci kereslet tehetetlensége.",
    "x_j = hx·x_j(−1) + (1−hx)·(−mu_x·(p_j − rer)).",
    "🟡 Közös, de az eltérő phi_j és p_j miatt a típuseltérést ERŐSÍTHETI.",
    "Az exportegyenlet AR(1) együtthatója.", "Export- és reálárfolyam-idősor.",
    "0,4–0,8 (JV 0,507)", JV_MOST, JV_KELL, "IGEN.", "hx")
add("Exportkereslet ár-rugalmassága", "Az export reakciója a relatív árra.",
    "Ugyanabban az exportegyenletben.",
    "🟡 Közös, de a típusonként eltérő exportkitettségen át differenciál.",
    "Az exportegyenlet relatívár-együtthatója (Armington-rugalmasság).",
    "Export, exportár, reálárfolyam idősor.",
    "0,5–2,0 (Armington-becslések; JV 0,534)", JV_MOST, JV_KELL, "IGEN.", "mu_x")
add("KKV-részesedés az exportszektor hazai köztes inputjában",
    "A hazai (KKV-) beszállítói input súlya az exportoldali határköltségben.",
    "mc_x_rel = (1−s_kkv)·(…) + s_kkv·mc_d − px; a shd_v inputja.",
    "🟡 A vertikális (beszállítói) csatorna erőssége. ⚠ A NÉV FÉLREVEZETŐ: ez NEM a KKV-k GDP-súlya.",
    "s_kkv = a hazai KKV-któl származó köztes input / az exportszektor összes köztes inputja.",
    "Input-output tábla vállalatméret szerinti bontással, VAGY hozzáadottérték-alapú kereskedelmi adat.",
    "Országspecifikus.",
    "❌ NEM a jelenlegi úton: az IO-alapú számítás HIBÁS (az összegzett köztes felhasználás a nemzeti számlák P2-jének csak 1,8–8,6%-a); a gyökérok nyitott, az IRÁNYT sem tudjuk (V02 visszavonva).",
    "OECD TiVA/ICIO magyar country note — MEGKERÜLI a hibás IO-számítást. De új, fogalmilag megfelelő vertikális-link módszertan kell hozzá.",
    "NEM.", "s_kkv")
add("Vertikális átgyűrűzési rugalmasság",
    "A hazai és az exportoldali határköltség közti ár-/mennyiségi átgyűrűzés erőssége.",
    "h_dx = xx − mu_vert·s_kkv·(mc_d − mc_x_rel).",
    "🟡 A beszállítói csatorna meredeksége.",
    "A beszállítói ár-átgyűrűzés regressziója: Δ(beszállítói ár) / Δ(vevői határköltség).",
    "Beszállító–vevő párosított árszintadat.",
    "❌ NEM TALÁLTUNK becslést a keresés során — ez a húszból az EGYETLEN, amiről a keresés után is azt mondjuk, hogy nincs mit megnézni.",
    "❌ NEM.", "Párosított beszállító–vevő tranzakciós adat (pl. ÁFA-tranzakciós adat).",
    "NEM — érzékenységi paraméterként kezelendő.", "mu_vert")
add("Típusok közti helyettesítési rugalmasság",
    "Mennyire helyettesíthető egymással a három vállalattípus terméke a hazai keresletben.",
    "d_j = y_d − eps_ces·p_j.",
    "🔴 Formálisan KÖZÖS, de az eltérő p_j miatt KRITIKUS szegmensdriver: az export-KKV kibocsátásának ELŐJELE fordul rajta ~2,3-nál (F02). Az aggregált GDP érzéketlen rá (0,008 pp sáv).",
    "eps_ces = markup/(markup−1) alakból: 20%-os markup → eps ≈ 6. Cégszintű markup-becslésből (De Loecker–Warzynski): markup = theta_input / (input-költséghányad).",
    "Cégszintű termelési adat (output, inputköltség) markup-becsléshez.",
    "6,0 az irodalmi konvenció (Laxton–Pesenti 2003 → GEM → EAGLE). ⚠ MAGYAR BECSLÉS LÉTEZIK: Dobrinsky–Kőrösi–Markov–Halpern (JCE 2006, 34(1):92–110).",
    "⚠ RÉSZBEN — a panel tartalmazza a markup-becsléshez kellő sorokat, DE FOGALMI CSÚSZÁS: nálunk az eps_ces a VÁLLALATTÍPUSOK közti helyettesítés, az irodalmi 6,0 viszont a TERMÉKVÁLTOZATOK közti. NEM UGYANAZ AZ OBJEKTUM.",
    "Elvi tisztázás előbb: mit helyettesít a fogyasztó — terméket vagy vállalattípust? Utána magyar markup-becslés plauzibilitási sávnak.",
    "⚠ RÉSZBEN — a 6,0 konvenció átvehető, de más objektumra vonatkozik; ezt ki kell mondani.",
    "eps_ces")

# --- sokk-perzisztenciák --------------------------------------------------
for nm, mi in [("rho_a", "technológiai"), ("rho_x", "exportkeresleti"),
               ("rho_c", "fogyasztási"), ("rho_w", "bér-"),
               ("rho_i", "beruházási"), ("rho_pr", "kockázati prémium"),
               ("rho_g", "kormányzati keresleti")]:
    add(f"{mi.capitalize()} sokk perzisztenciája", SOKK,
        "A megfelelő AR(1) sokkfolyamat tehetetlensége.",
        "⚪ A determinisztikus euró-szcenárióban NEM aktív. Sztochasztikus futásban a sokk lecsengésének sebessége.",
        SOKK_KEPLET, JV_ADAT, "0,4–0,9 (tipikus DSGE-tartomány)", JV_MOST, JV_KELL, "IGEN.", nm)
add("Exportár/markup sokk perzisztenciája", SOKK,
    "Az e_mx_ar AR(1) folyamat tehetetlensége.",
    "⚠ KIVÉTEL a sokkok közül: a rho_mx az L-TÍPUS ár-Phillips-sokkjának tartósságát szabja (a pi_E/pi_D nyers sokkot kap, a pi_L AR-folyamatot). Sztochasztikus futásban ez NÉMÁN eltolhatná a KKV/nagyvállalat összevetést — füstteszt-őr fogja el.",
    SOKK_KEPLET, JV_ADAT, "0,2–0,6 (JV 0,318)", JV_MOST, JV_KELL, "IGEN.", "rho_mx")

# --- származtatottak ------------------------------------------------------
add("Ár-Phillips-görbe meredeksége", SZARM,
    "A határköltség hatása az inflációra.", "🟡 Aktív és valóban származtatott.",
    "lam_p = (1−xi_p)(1−beta·xi_p)/xi_p", "— (az xi_p és beta örökölt bizonytalansága)",
    "—", "SZÁRMAZTATOTT.", "—", "—", "lam_p")
add("Export-Phillips meredekség ⚫ HALOTT", SZARM,
    "⚫ Kiszámolódik, de a v09 model blokkjában NEM szerepel.", "SEMMIRE.",
    "lam_x = (1−xi_x)(1−beta·xi_x)/xi_x", "—", "—", "TÁRGYTALAN.", "—", "—", "lam_x")
add("Bér-Phillips-görbe meredeksége", SZARM,
    "A bérrés hatása a bérinflációra.", "⚪/🟡 Aktív és származtatott.",
    "lam_w = (1−xi_w)(1−beta·xi_w)/(xi_w·(1+theta_w·fii))", "—", "—",
    "SZÁRMAZTATOTT.", "—", "—", "lam_w")
for t in "EDL":
    add("Hazai értékesítési súly", SZARM,
        "A hazai határköltség és a relatívár-normalizálás súlya.",
        "🔵/🟡 ⚠ ÖRÖKÖLT BIZONYTALANSÁG: az om_j (feltételes) és a phi_j (feltételes) hibáját viszi tovább.",
        "wd_j = om_j(1−phi_j) / Σ_k om_k(1−phi_k)", "—", "—",
        "SZÁRMAZTATOTT — az inputjai határozzák meg.", "—", "—", f"wd_{t}")
    add("Export értékesítési súly", SZARM,
        "Az exportár és -mennyiség aggregációs súlya.",
        "🔵/🟡 ⚠ ÖRÖKÖLT: om_j és phi_j. Az exportoldali KKV/nagyvállalat összetételt ÉRDEMBEN mozgatja.",
        "wx_j = om_j·phi_j / Σ_k om_k·phi_k", "—", "—",
        "SZÁRMAZTATOTT.", "—", "—", f"wx_{t}")
    add("Import-súly", SZARM, "Az aggregált importkereslet súlya.",
        "🔵/🟡 ⚠ ÖRÖKÖLT: om_j (feltételes) és aa_j (átvett, pótolandó).",
        "shm_j = om_j(1−aa_j) / Σ_k om_k(1−aa_k)", "—", "—",
        "SZÁRMAZTATOTT.", "—", "—", f"shm_{t}")
for nm, cim, kep in [("shd_v", "Vertikális komponens súlya", "shd_v = s_kkv · 0,60"),
                     ("shd_c", "Fogyasztási komponens súlya", "shd_c = 0,55(1−shd_v)/0,82"),
                     ("shd_i", "Beruházási komponens súlya", "shd_i = 0,15(1−shd_v)/0,82"),
                     ("shd_g", "Kormányzati komponens súlya", "shd_g = 0,12(1−shd_v)/0,82")]:
    add(cim, SZARM, "A hazai kereslet összetétele.",
        "🔵/🟡 ⚠ ÖRÖKÖLT: a HORGONYZATLAN s_kkv bizonytalanságát viszi. ⚠ A képletben szereplő 0,60/0,55/0,15/0,12/0,82 konstansok NINCSENEK paraméterként deklarálva — a 91-es auditon kívül maradnak.",
        kep, "—", "—", "SZÁRMAZTATOTT.", "—", "—", nm)

FEJ = ["Jele", "Neve", "Jelentése", "Szerepe a modellben",
       "Mit határoz meg / mire hat", "Hogyan számolható ki (képlet + szóban)",
       "Adatigény", "Szakirodalmi érték és intervallum",
       "Megcsinálható a jelenlegi adatból?", "Ha nem, milyen adat kell",
       "Konvencionális? (átvehető-e)"]


def main() -> None:
    reg = pd.read_csv(REG / "parameterek.csv")
    ert = pd.read_csv(REG / "_params_dump.csv")
    df = reg.merge(ert, on="parameter", how="left")

    hiany = [p for p in df["parameter"] if p not in P]
    if hiany:
        raise SystemExit(f"Hiányzó kalibrációs leírás ({len(hiany)}): {hiany}")

    sorok = []
    for _, r in df.iterrows():
        n = r["parameter"]
        v = P[n]
        sorok.append({
            "Jele": n, "Neve": v[0], "Jelentése": v[1], "Szerepe a modellben": v[2],
            "Mit határoz meg / mire hat": v[3],
            "Hogyan számolható ki (képlet + szóban)": v[4],
            "Adatigény": v[5], "Szakirodalmi érték és intervallum": v[6],
            "Megcsinálható a jelenlegi adatból?": v[7],
            "Ha nem, milyen adat kell": v[8], "Konvencionális? (átvehető-e)": v[9],
            "Jelenlegi érték (OPTEN=0)": r["ertek_OPTEN0"],
            "Kategória": r["kategoria"], "Státusz": r["statusz"],
            "Hatás": r["hatas"], "Kapcsoló": r["kapcsolo"] if isinstance(r["kapcsolo"], str) else "",
        })
    out = pd.DataFrame(sorok)
    oszl = FEJ + ["Jelenlegi érték (OPTEN=0)", "Kategória", "Státusz", "Hatás", "Kapcsoló"]
    out = out[oszl]

    KI.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(KI, engine="openpyxl") as w:
        out.to_excel(w, sheet_name="Kalibracios terv", index=False)
        _osszegzo(w, len(out))
    _formaz(out)
    print(f"KIÍRVA: {KI}  ({len(out)} paraméter)")


def _osszegzo(writer, n: int) -> None:
    """Összegző lap — a darabszámok KÉPLETTEL, hogy a fő lappal együtt mozogjon."""
    cimek = ["Kategória / státusz", "Darab"]
    kulcsok = [("A — saját adat (Opten)", 'COUNTIF(\'Kalibracios terv\'!M:M,"A")'),
               ("B — KSH-ból pótolható", 'COUNTIF(\'Kalibracios terv\'!M:M,"B")'),
               ("C — JV / irodalmi konvenció", 'COUNTIF(\'Kalibracios terv\'!M:M,"C")'),
               ("D — nem azonosított", 'COUNTIF(\'Kalibracios terv\'!M:M,"D")'),
               ("E — származtatott / technikai", 'COUNTIF(\'Kalibracios terv\'!M:M,"E")'),
               ("", ""),
               ("horgonyzott", 'COUNTIF(\'Kalibracios terv\'!N:N,"horgonyzott")'),
               ("feltételes", 'COUNTIF(\'Kalibracios terv\'!N:N,"feltételes")'),
               ("pótolandó", 'COUNTIF(\'Kalibracios terv\'!N:N,"pótolandó")'),
               ("horgonyzatlan", 'COUNTIF(\'Kalibracios terv\'!N:N,"horgonyzatlan")'),
               ("származtatott", 'COUNTIF(\'Kalibracios terv\'!N:N,"származtatott")'),
               ("", ""),
               ("HALOTT paraméter", 'COUNTIF(\'Kalibracios terv\'!O:O,"HALOTT*")'),
               ("örökölt bizonytalanságú", 'COUNTIF(\'Kalibracios terv\'!O:O,"örökölt*")'),
               ("", ""),
               ("ÖSSZESEN", f"COUNTA('Kalibracios terv'!A2:A{n + 1})")]
    ws = writer.book.create_sheet("Osszegzo")
    ws.append(cimek)
    for cim, kep in kulcsok:
        ws.append([cim, f"={kep}" if kep else ""])
    ws["A1"].font = Font(name="Arial", bold=True)
    ws["B1"].font = Font(name="Arial", bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 10
    for row in ws.iter_rows():
        for c in row:
            if not c.font.bold:
                c.font = Font(name="Arial")
    ws["A19"] = ("A darabszámok KÉPLETTEL számolódnak a 'Kalibracios terv' lapról, "
                 "tehát együtt mozognak vele.")
    ws["A19"].font = Font(name="Arial", italic=True, size=9)
    ws["A20"] = "Forrás: docs/regiszter/parameterek.csv + _params_dump.csv (a modellből)."
    ws["A20"].font = Font(name="Arial", italic=True, size=9)
    ws["A21"] = ("MEGJEGYZES: a B oszlop cellai KEPLETEK. Ha a megnyitaskor uresnek "
                 "latszanak, nyomj Ctrl+Alt+F9-et (ujraszamolas) - a generalo kornyezetben "
                 "nem volt LibreOffice, ezert nincs elmentett gyorsitotar-ertekuk.")
    ws["A21"].font = Font(name="Arial", italic=True, size=9, color="C00000")


def _formaz(out: pd.DataFrame) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(KI)
    ws = wb["Kalibracios terv"]
    SZEL = {"A": 15, "B": 34, "C": 42, "D": 42, "E": 52, "F": 56, "G": 34,
            "H": 46, "I": 46, "J": 42, "K": 34, "L": 14, "M": 10, "N": 14,
            "O": 30, "P": 12}
    for k, v in SZEL.items():
        ws.column_dimensions[k].width = v
    fej_kit = PatternFill("solid", fgColor="1F3864")
    keret = Side(style="thin", color="BFBFBF")
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = fej_kit
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = Border(bottom=keret)
    ws.row_dimensions[1].height = 46
    # a D kategoria (nem azonositott) sorai sargan -- ezekkel kell dolgozni
    sarga = PatternFill("solid", fgColor="FFF2CC")
    for i, kat in enumerate(out["Kategória"], start=2):
        for c in ws[i]:
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = Border(bottom=keret)
            if kat == "D":
                c.fill = sarga
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(KI)


if __name__ == "__main__":
    main()
